from django.contrib import admin
from .models import Sclass, Student,Fee_Category, Fee
import csv
import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
from django.utils.translation import gettext_lazy as _
#from daterangefilter.filters import PastDateRangeFilter, FutureDateRangeFilter
#from daterangefilter.filters import DateRangeFilter
from django_filters import ModelChoiceFilter

from rangefilter.filters import (
    DateRangeFilterBuilder,
    DateTimeRangeFilterBuilder,
    NumericRangeFilterBuilder,
    DateRangeQuickSelectListFilterBuilder,
)
from django.db import models
from django.contrib import admin
from django.forms import ModelChoiceField
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from .models import Fee, Student, FeeSet
from django.utils.html import format_html
from django import forms
from django_select2.forms import ModelSelect2Widget

# from .faker import generate_fake_students

# generate_fake_students()
class SclassAdmin(admin.ModelAdmin):
    list_display = ('name','created_date',)

class FeeCategoryAdmin(admin.ModelAdmin):
    search_fields = ['name']
    list_display = ('name','created_date',)

# class FeeAdmin(admin.ModelAdmin):
#     list_display = ('student','fees_category',)

def export_to_pdf(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="students.pdf"'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    students = []

    # Title
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1  # Center alignment
    title_style.fontSize = 20  # Font size
    title_text = "List of Students"  # Title text

    # Define the title
    title_content = Paragraph(title_text, title_style)
    students.append([title_content])

    # Add space after the title
    students.append([Spacer(1, 12)])

    # Define the table data
    for student in queryset:
        students.append([student.name, student.dob, student.section, student.roll_number, student.gender, student.blood_group, student.parents_phone_number, student.admission_id])

    # Create the table
    table = Table(data=students)

    # Style the table
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)])

    table.setStyle(style)

    # Add title and table to the PDF
    doc.build([title_content, Spacer(1, 12), table])
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

export_to_pdf.short_description = 'Export to PDF'

def export_to_excel(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Date of Birth', 'Section', 'Roll Number', 'Gender', 'Blood Group', 'Parents Phone Number', 'Admission ID'])
    
    for student in queryset:
        writer.writerow([student.name, student.dob, student.section, student.roll_number, student.gender, student.blood_group, student.parents_phone_number, student.admission_id])
    
    return response

export_to_excel.short_description = 'Export to Excel'

class SclassFilter(admin.SimpleListFilter):
    title = _('Sclass')
    parameter_name = 'sclass'

    def lookups(self, request, model_admin):
        return Sclass.objects.values_list('id', 'name')

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(section=self.value())

class StudentAdmin(admin.ModelAdmin):
    list_display = ('name', 'payment_times', 'total_due_display', 'due_explain_display', 'section', 'roll_number', 'gender', 'parents_phone_number', 'admission_id')
    list_filter = [("created_date", DateRangeFilterBuilder()),
                   ("roll_number", NumericRangeFilterBuilder()),
                   (SclassFilter),
                   'gender', 
                   'blood_group']
    search_fields = ('name', 'roll_number', 'admission_id')  
    def total_due_display(self, obj):
        total_due_info = obj.total_due
        color = total_due_info['color']
        amount = total_due_info['amount']
        return format_html('<span style="color:{}">{}</span>', color, amount)

    total_due_display.allow_tags = True
    total_due_display.short_description = 'Total Due'

    def due_explain_display(self, obj):
        return format_html(obj.due_explain)

    due_explain_display.allow_tags = True
    due_explain_display.short_description = 'Due Explanation'

    actions = [export_to_excel, export_to_pdf]

admin.site.register(Student, StudentAdmin)
# Register your models with custom ModelAdmin classes
admin.site.register(Sclass, SclassAdmin)
admin.site.register(Fee_Category, FeeCategoryAdmin)

from django import forms

class FeeAdminForm(forms.ModelForm):
    class Meta:
        model = Fee
        fields = '__all__'

    class Media:
        js = ('admin/js/fee_admin.js',)

class StatusFilter(admin.SimpleListFilter):
    title = _('Status')
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('Paid', _('Paid')),
            ('Due', _('Due')),
        )

    def queryset(self, request, queryset):
        if self.value() == 'Paid':
            return queryset.filter(partial_payment__gte=models.F('fees_amount'))
        elif self.value() == 'Due':
            return queryset.exclude(partial_payment__gte=models.F('fees_amount'))
        

import openpyxl
from reportlab.pdfgen import canvas
class FeeAdmin(admin.ModelAdmin):
    form = FeeAdminForm
    list_display = ['student', 'fee_category', 'fees_amount', 'partial_payment', 'get_status', 'due_payment', 'voucher_id']
    autocomplete_fields = ['student']
    search_fields = ['student__name',]
    list_filter = [("created_date", DateRangeFilterBuilder()),StatusFilter,
                   'fee_category', ]

    def get_status(self, obj):
        # Get the status of the fee
        status = obj.status
        # Determine the color based on the status
        if status == 'Paid':
            color = 'green'
        else:
            color = 'red'
        # Return the status with HTML style for color
        return format_html('<span style="color:{}">{}</span>', color, status)
    get_status.short_description = 'Status'  # Set the column header name
    
    def download_pdf(self, request, queryset):
        # Create a PDF document
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="fees.pdf"'
        
        # Define the table data
        table_data = []
        table_data.append(['Student', 'Fee Category', 'Fees Amount', 'Partial Payment', 'Total Paid', 'Voucher ID'])
        total_fees = 0
        total_partial_payments = 0
        total_due = 0
        for fee in queryset:
            total_paid = fee.fees_amount - fee.partial_payment
            table_data.append([fee.student, fee.fee_category, fee.fees_amount, fee.partial_payment, total_paid, fee.voucher_id])
            total_fees += fee.fees_amount
            total_partial_payments += fee.partial_payment
            total_due += fee.fees_amount - fee.partial_payment

        # Add footer rows for total fee amount, total partial payments, and total due amount
        table_data.append(['', '', f'Total Fees: {total_fees}', f'Total Partial Payments: {total_partial_payments}', f'Total Due: {total_due}', ''])

        # Create the PDF document
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []

        # Create the table
        table = Table(table_data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])
        table.setStyle(style)

        # Add the table to the document
        elements.append(table)
        doc.build(elements)

        return response
    download_pdf.short_description = _('Download PDF')

    def download_excel(self, request, queryset):
        # Create a new Excel workbook
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="fees.xlsx"'

        # Create a workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fees"

        # Define the table headers
        headers = ['Student', 'Fee Category', 'Fees Amount', 'Partial Payment', 'Total Paid', 'Voucher ID']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Add data rows and calculate total fee amount, total partial payments, and total due amount
        total_fees = 0
        total_partial_payments = 0
        total_due = 0
        for row_num, fee in enumerate(queryset, 2):
            total_paid = fee.fees_amount - fee.partial_payment
            ws.cell(row=row_num, column=1, value=str(fee.student))
            ws.cell(row=row_num, column=2, value=str(fee.fee_category))
            ws.cell(row=row_num, column=3, value=float(fee.fees_amount))
            ws.cell(row=row_num, column=4, value=float(fee.partial_payment))
            ws.cell(row=row_num, column=5, value=float(total_paid))
            ws.cell(row=row_num, column=6, value=str(fee.voucher_id))
            total_fees += float(fee.fees_amount)
            total_partial_payments += float(fee.partial_payment)
            total_due += float(fee.fees_amount) - float(fee.partial_payment)

        # Add footer rows for total fee amount, total partial payments, and total due amount
        total_row = len(queryset) + 2
        ws.cell(row=total_row, column=2, value='Total Fees:')
        ws.cell(row=total_row, column=3, value=total_fees)
        ws.cell(row=total_row, column=4, value='Total Partial Payments:')
        ws.cell(row=total_row, column=5, value=total_partial_payments)
        ws.cell(row=total_row, column=6, value='Total Due:')
        ws.cell(row=total_row, column=7, value=total_due)

        # Save the workbook
        wb.save(response)

        return response
    download_excel.short_description = _('Download Excel')

    # Add the custom actions to the admin actions dropdown
    actions = ['download_pdf', 'download_excel']
    

admin.site.register(Fee, FeeAdmin)



class FeeSetAdmin(admin.ModelAdmin):
    list_display = ['fee_category', 'class_category', 'year', 'fee_amount', 'created_date', 'modified_date']
    search_fields = ['fee_category__name', 'class_category__name']
    list_filter = ['created_date', 'modified_date']
    

admin.site.register(FeeSet, FeeSetAdmin)

admin.site.site_header = "Darsana DS Fazil Madrasah"
admin.site.site_title = "Darsana DS Fazil MadrasahAdmin Portal"
admin.site.index_title = "Welcome to Darsana DS Fazil Madrasah Admin Portal"


# from .faker_fees import generate_fake_fees
# generate_fake_fees()

